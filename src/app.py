import sys
from PyQt5.QtWidgets import (QApplication, QMainWindow, QPushButton, QVBoxLayout, QWidget,
                             QFileDialog, QLabel, QListWidget, QMessageBox, QProgressBar)
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import re
import os


def combine_excel_sheets(file_paths, output_file, progress_callback):
 
    with pd.ExcelWriter(output_file, mode='w') as writer:
        pd.DataFrame().to_excel(writer, sheet_name='Test', index=False)
 
    def load_all_worksheets(file_paths):
        all_dataframes = {}
        for file_path in file_paths:
            try:
                xls = pd.ExcelFile(file_path)
                base_name = os.path.basename(file_path).split('.')[0]
                for sheet_name in xls.sheet_names:
                    key = f"{base_name} - {sheet_name}"
                    all_dataframes[key] = xls.parse(sheet_name)
            except Exception as e:
                missed_files.append(file_path)
                print(f"Error reading the file {file_path}: {e}")
        return all_dataframes
 
    missed_files = []
    all_sheets_data = load_all_worksheets(file_paths)
 
    total_files = len(all_sheets_data)+33
    processed_files = 0

    print(f'Input: {len(file_paths)} files')
    print(f'No errors: {len(file_paths)-len(missed_files)} files')
    print(f'Files with errors: {missed_files}')
 
    cad_cols = ['Отечественное/Зарубежное', 'Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'Статус применимости', 'Наименование ПО', 'Разработчик ПО', 'На скольки АРМ установлено',
                'На какой базе данных (отечественной/зарубежной)', 'Количество официальных лицензий', 'Количество прочих копий лицензий']
    ecad_cols = ['Отечественное/Зарубежное', 'Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'Статус применимости', 'Наименование ПО', 'Разработчик ПО', 'На скольки АРМ установлено',
                'На какой базе данных (отечественной/зарубежной)', 'Количество официальных лицензий', 'Количество прочих копий лицензий']
    cae_cols = ['Отечественное/Зарубежное', 'Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'Статус применимости', 'Наименование ПО', 'Разработчик ПО', 'На скольки АРМ установлено',
                'На какой базе данных (отечественной/зарубежной)', 'Количество официальных лицензий', 'Количество прочих копий лицензий', 'Применяется ли для проведения суперкомпьютерных вычислений']
    capp_cols = ['Отечественное/Зарубежное', 'Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'Статус применимости', 'Наименование ПО', 'Разработчик ПО', 'На скольки АРМ установлено',
                'На какой базе данных (отечественной/зарубежной)', 'Количество официальных лицензий', 'Количество прочих копий лицензий']
    cam_cols = ['Отечественное/Зарубежное', 'Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'Статус применимости', 'Наименование ПО', 'Разработчик ПО', 'На скольки АРМ установлено',
                'На какой базе данных (отечественной/зарубежной)', 'Количество официальных лицензий', 'Количество прочих копий лицензий']
    pdm_cols = ['Отечественное/Зарубежное', 'Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'Статус применимости', 'Наименование ПО', 'Разработчик ПО', 'На скольки АРМ установлено',
                'На какой базе данных (отечественной/зарубежной)', 'Количество официальных лицензий', 'Количество прочих копий лицензий']
    erp_cols = ['Отечественное/Зарубежное', 'Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'Статус применимости', 'Наименование ПО', 'Разработчик ПО', 'На скольки АРМ установлено',
                'На какой базе данных (отечественной/зарубежной)', 'Количество официальных лицензий', 'Количество прочих копий лицензий']
    subu_cols = ['Отечественное/Зарубежное', 'Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'Статус применимости', 'Наименование ПО', 'Разработчик ПО', 'На скольки АРМ установлено',
                'На какой базе данных (отечественной/зарубежной)', 'Количество официальных лицензий', 'Количество прочих копий лицензий']
    sb_cols = ['Отечественное/Зарубежное', 'Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'Статус применимости', 'Наименование ПО', 'Разработчик ПО', 'На скольки АРМ установлено',
                'На какой базе данных (отечественной/зарубежной)', 'Количество официальных лицензий', 'Количество прочих копий лицензий']
    supr_cols = ['Отечественное/Зарубежное', 'Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'Статус применимости', 'Наименование ПО', 'Разработчик ПО', 'На скольки АРМ установлено',
                'На какой базе данных (отечественной/зарубежной)', 'Количество официальных лицензий', 'Количество прочих копий лицензий']
    sup_cols = ['Отечественное/Зарубежное', 'Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'Статус применимости', 'Наименование ПО', 'Разработчик ПО', 'На скольки АРМ установлено',
                'На какой базе данных (отечественной/зарубежной)', 'Количество официальных лицензий', 'Количество прочих копий лицензий']
    mrp2_cols = ['Отечественное/Зарубежное', 'Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'Статус применимости', 'Наименование ПО', 'Разработчик ПО', 'На скольки АРМ установлено',
                'На какой базе данных (отечественной/зарубежной)', 'Количество официальных лицензий', 'Количество прочих копий лицензий']
    ils_cols = ['Отечественное/Зарубежное', 'Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'Статус применимости', 'Наименование ПО', 'Разработчик ПО', 'На скольки АРМ установлено',
                'На какой базе данных (отечественной/зарубежной)', 'Количество официальных лицензий', 'Количество прочих копий лицензий']
    iatr_cols = ['Отечественное/Зарубежное', 'Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'Статус применимости', 'Наименование ПО', 'Разработчик ПО', 'На скольки АРМ установлено',
                'На какой базе данных (отечественной/зарубежной)', 'Количество официальных лицензий', 'Количество прочих копий лицензий']
    mdm_cols = ['Отечественное/Зарубежное', 'Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'Количество справочников в MDM-системе', 'Согласованы ли MDM-системы с MDM интегрированной структуры',
                'Статус применимости MDM', 'Наименование ПО', 'Разработчик ПО', 'На скольки АРМ установлено', 'На какой базе данных (отечественной/зарубежной)',
                'Количество официальных лицензий', 'Количество прочих копий лицензий', 'Количество справочников вне MDM-систем (внутри локальных ПО всех классов задач)',
                'На скольки АРМ имеются справочники', 'Количество справочников, синхронизированных между ПО различных классов', 'Количество справочников, синхронизированных между ДЗО']
    sad_cols = ['Отечественное/Зарубежное', 'Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'Статус применимости', 'Наименование ПО', 'Разработчик ПО', 'На скольки АРМ установлено',
                'На какой базе данных (отечественной/зарубежной)', 'Количество официальных лицензий', 'Количество прочих копий лицензий']
    eam_cols = ['Отечественное/Зарубежное', 'Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'Статус применимости', 'Наименование ПО', 'Разработчик ПО', 'На скольки АРМ установлено',
                'На какой базе данных (отечественной/зарубежной)', 'Количество официальных лицензий', 'Количество прочих копий лицензий']
    reglamenty_cols = ['Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'On-line автоматически', 'Дискретно автоматизировано', 'Дискретно вручную', 'Отсутствуют автоматизированные информационные системы']
    kommunikazii_cols = ['', '', '', '', '', '']
    cody_cols = ['Наименование ИС/ДЗО (20)', 'Да, свой (20)', 'Да, управляющей компании (20)', 'Да, арендую ЦОД (20)', 'Нет, но нужен (20)', 'Нет, не нужен (20)',
                'Наименование ИС/ДЗО (20.1)', 'автоматическое (20.1)', 'автоматизированное (20.1)', 'ручное (20.1)',
                'Наименование ИС/ДЗО (20.2)', 'открытые и дсп (20.2)', 'секретно и выше (20.2)', 'присутствуют оба типа (20.2)', 'отсутствует категорирование (20.2)']
    skt_cols = ['Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'Применяются ли суперкомпьютерные технологии', 'Потребность в аренде вычислительных ресурсов', 'Стадия применения',
                'Необходимая производительность, Тфлоп/с для открытых и дсп данных', 'Необходимая производительность, Тфлоп/с для секретных данных и выше',
                'Применяемое ИС ПО для проведения суперкомпьютерных вычислений',
                'Производительность собственного суперкомпьютера организации (суммарная производительность нескольких суперкомпьютеров), Тфлопс']
    obshesistemnoe_po_cols = ['Отечественное/Зарубежное', 'Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'Класс ПО', 'Наименование ПО', 'Разработчик ПО', 'На скольки АРМ установлено', 'Количество официальных лицензий',
                'Количество прочих копий лицензий', ]
    intergracia_oborudovaniya_cols = ['Наименование ИС/ДЗО', 'ОКПО ИС/ДЗО', 'Механообработка с ЧПУ', 'Гальваника', 'Сварка', 'Литейка', 'Штамповка', 'Лазеры', 'Печатные платы', 'Аддитивные технологии']
    sistemy_monitoringa_cols = ['Наименование ИС/ДЗО', 'Металлургическое производство (В контуре)', 'Металлургическое производство (Вне контура)',
                'Производство оптических материалов и оптических сред (В контуре)', 'Производство оптических материалов и оптических сред (Вне контура)',
                'Литейное производство (В контуре)', 'Литейное производство (Вне контура)',
                'Кузнечно-прессовое производство (В контуре)', 'Кузнечно-прессовое производство (Вне контура)',
                'Механообрабатывающее производство (В контуре)', 'Механообрабатывающее производство (Вне контура)',
                'Оптическое производство (В контуре)', 'Оптическое производство (Вне контура)',
                'Аддитивное производство (В контуре)', 'Аддитивное производство (Вне контура)',
                'Химическое производство (В контуре)', 'Химическое производство (Вне контура)',
                'Гальваническое производство (В контуре)', 'Гальваническое производство (Вне контура)',
                'Химикотермическое производство, включая нанесение защитных покрытий (В контуре)', 'Химикотермическое производство, включая нанесение защитных покрытий (Вне контура)',
                'Производство и переработка пластмасс, полимерных материалов, резинотехнических изделий (В контуре)', 'Производство и переработка пластмасс, полимерных материалов, резинотехнических изделий (Вне контура)',
                'Изоляционное производство (В контуре)', 'Изоляционное производство (Вне контура)',
                'Производство изделий из керамики и ферритов (В контуре)', 'Производство изделий из керамики и ферритов (Вне контура)',
                'Производство композиционных материалов, конструкций и изделий из них (В контуре)', 'Производство композиционных материалов, конструкций и изделий из них (Вне контура)',
                'Производство электронных компонентов (В контуре)', 'Производство электронных компонентов (Вне контура)',
                'Производство оптоэлектронных компонентов (В контуре)', 'Производство оптоэлектронных компонентов (Вне контура)',
                'Производство и монтаж печатных плат (В контуре)', 'Производство и монтаж печатных плат (Вне контура)',
                'Сборочно-сварочное производство (В контуре)', 'Сборочно-сварочное производство (Вне контура)',
                'Сборочно-монтажное производство (В контуре)', 'Сборочно-монтажное производство (Вне контура)',
                'Деревообрабатывающее и тарное производство (В контуре)', 'Деревообрабатывающее и тарное производство (Вне контура)',
                'Инструментальное производство (В контуре)', 'Инструментальное производство (Вне контура)',
                'Производство нестандартного оборудования и спецоснастки (В контуре)', 'Производство нестандартного оборудования и спецоснастки (Вне контура)',
                'Заводские и полигонные испытания (В контуре)', 'Заводские и полигонные испытания (Вне контура)',
                'Контрольно-метрологические работы (В контуре)', 'Контрольно-метрологические работы (Вне контура)',
                'Прочие виды производства (В контуре)', 'Прочие виды производства (Вне контура)']
    standarty_cols = ['Наименование ИС/ДЗО', 'Применение ПО по классам задач при проектировании, производстве и эксплуатации изделий', 'Легализацию бизнес-процессов реализованных в различных классах задач',
                'Легализацию отчетных форм', 'Легализацию BI аналитики']
    bi_sistemy_cols = ['Наименование ИС/ДЗО', 'PDM', 'MES', 'Объемно-календарное планирование', 'Производство', 'Склад', 'Снабжение', 'Бюджетирование', 'Бухгалтерия', 'ИЭТР', 'Мониторинг загрузки оборудования']
    ORD_cols = ['Наименование ИС/ДЗО', 'Установленное внутренним регламентом плановое количество АРМ с сертифицированной ЭЦП', 'Фактическое количество АРМ с сертифицированной ЭЦП',
                'Установленное внутренним регламентом плановое количество АРМ с ЭЦП по внутренним стандартам', 'Фактическое количество АРМ с ЭЦП по внутренним стандартам',
                'Установленное внутренним регламентом плановое количество АРМ без ЭЦП', 'Фактическое количество АРМ без ЭЦП']
    kd_cols = ['Наименование ИС/ДЗО', 'Автоматизация задач в части обеспечения связи договоров на исполнение Госконтрактов с расчетными счетами предприятия, используемыми для осуществления взаиморасчетов по этим Госконтрактам',
                'Автоматизация задач в части расчета заработной платы, НДФЛ и налогов на ФОТ в разрезе Госконтрактов',
                'Автоматизация задач в части передачи на электронные сервисы банков подтверждающих платежных документов (сканобразов) для последующего контроля банками движения денежных средств по специальным банковским счетам',
                'Автоматизации задач в части подготовки отчетности по исполнению ГОЗ для отслеживания платежей по отгрузкам/ поставкам по Госконтрактам',
                'Автоматизация управления госконтрактами', 'Автоматизация управления финансовыми ресурсами (бюджетирование, платежный календарь)',
                'Автоматизация расчета плановой себестоимости продукции и плановых экономических показателей',
                'Автоматизации расчета фактической себестоимости продукции и экономических показателей, план/факт анализ',
                'Автоматизации управления инвестиционными проектами и программами']
    mzk_cols = ['Наименование ИС/ДЗО', 'Единые стандарты обмена файлами, единое интегрированное информационное пространство проектирования и информационного обмена с предприятиями-партнерами для большинства продуктов',
                'Автоматизированы решения задач по построению графиков кооперации', 'Автоматизированы решения задач по анализу загрузки ресурсов предприятий, входящих в цепочку кооперации',
                'Автоматизированы решения задач по контролю хода выполнения кооперационного производства', 'Автоматизированы решения задач по учету выполнения контрактов в рамках заданного графика кооперации',
                'Автоматизированы решения задач в сфере организации взаимоотношений с поставщиками в рамках заданной сети кооперации', 'Автоматизированы решения задач по выбору оптимальных поставщиков',
                'Автоматизированы задачи в сфере организации взаимоотношений с головным предприятием / холдингом в рамках заданной сети кооперации']
    kadry_1_cols = ['Наименование ИС/ДЗО', 'CAD (не учитывается ECAD)', 'ECAD (без учета механических CAD)', 'CAE', 'CAPP', 'CAM', 'PDM', 'ERP', 'Системы управления бухгалтерским учетом',
                'Системы бюджетирования', 'Системы управления проектами', 'Системы управления персоналом', 'MRP-II', 'ILS', 'MDM', 'СЭД',
                'Практические занятия с персоналом по правилам эксплуатации системы защиты автоматизированной (информационной) системы и отдельных средств защиты информации',
                'Теоретическое обучение персонала правилам эксплуатации системы защиты автоматизированной (информационной) системы и отдельных средств защиты информации']
    kadry_2_cols = ['Наименование ИС/ДЗО', 'Наличие системы обучения', 'План (раз/год)', 'Факт (раз/год)']
    bim_cols = ['Наименование ИС/ДЗО', 'Применяется ли BIM-система', 'Наименование ПО', 'Цех,производство', 'Участок', 'Станок', 'Энергетическое оборудование', 'Системы складов',
                'Детали, полуфабрикаты', 'Техпроцессы', 'Работники']
    ib_cols = ['Наименование ИС/ДЗО', 'Наличие на предприятии назначенных должностных лиц, ответственных за организацию и контроль состояния защиты информации',
                'Наличие на предприятии постоянно-действующих технических комиссий (возможно, экспертные комиссии) по определению уровня конфиденциальности информации, образующейся при обработке сведений в автоматизированных (информационных) системах, по совокупности',
                'Наличие на предприятии постоянно-действующих комиссий или иных коллегиальных органов, управляющих и регламентирующих закупку и разрешение к использованию в автоматизированных (информационных) системах предприятия программных изделий и средств защиты информации',
                'Информирование и оценка осведомленности персонала об угрозах безопасности информации, о правилах эксплуатации системы защиты автоматизированной (информационной) системы и отдельных средств защиты информации',
                'Управление (администрирование) системой защиты информации автоматизированной (информационной) системы', 'Регламенты выявления инцидентов и реагирования на них',
                'Управление конфигурацией аттестованной автоматизированной (информационной) системы и ее системы защиты информации',
                'Контроль (мониторинг) за обеспечением уровня защищенности информации, содержащейся в автоматизированной (информационной) системе',
                'Обеспечение защиты информации при выводе из эксплуатации аттестованной автоматизированной (информационной) системы или после принятия решения об окончании обработки информации',
                'Политика обеспечения технической защиты информации на предприятии',
                'Руководство по защите информации предприятия',
                'План мероприятий по обеспечению технической защиты информации на предприятии',
                'Система менеджмента информационной безопасности на предприятии (часть общей системы менеджмента, основанная на использовании методов оценки рисков для разработки, внедрения, функционирования, мониторинга, анализа, поддержки и улучшения информационной безопасности).',
                'Менеджмент риска информационной безопасности (скоординированные действия по руководству и управлению предприятием в отношении риска ИБ с целью его минимизации)',
                'Меры по идентификации и аутентификации субъектов и объектов доступа в автоматизированных (информационных) системах',
                'Меры по управлению доступом субъектов к объектам доступа в автоматизированных (информационных) системах',
                'Меры по ограничению программной среды автоматизированных (информационных) систем',
                'Меры по защите машинных носителей информации (средств хранения информации, съемных машинных носителей информации)',
                'Меры по регистрации событий безопасности в автоматизированной (информационной) системе',
                'Меры по антивирусной защите в автоматизированной (информационной) системе',
                'Меры по обнаружению (предотвращению) вторжений в автоматизированную (информационную) систему',
                'Меры по контролю (анализу) защищенности информации в автоматизированной (информационной) системе',
                'Меры по защите среды виртуализации в автоматизированной (информационной) системе',
                'Меры по защите информации при ее передаче по каналам связи, защите компонентов автоматизированной (информационной) системы',
                'Меры по защите технических средств автоматизированной (информационной) системы']
 
    cad_df1 = pd.DataFrame()
    cad_df2 = pd.DataFrame()
    ecad_df1 = pd.DataFrame()
    ecad_df2 = pd.DataFrame()
    cae_df1 = pd.DataFrame()
    cae_df2 = pd.DataFrame()
    capp_df1 = pd.DataFrame()
    capp_df2 = pd.DataFrame()
    cam_df1 = pd.DataFrame()
    cam_df2 = pd.DataFrame()
    pdm_df1 = pd.DataFrame()
    pdm_df2 = pd.DataFrame()
    erp_df1 = pd.DataFrame()
    erp_df2 = pd.DataFrame()
    subu_df1 = pd.DataFrame()
    subu_df2 = pd.DataFrame()
    sb_df1 = pd.DataFrame()
    sb_df2 = pd.DataFrame()
    supr_df1 = pd.DataFrame()
    supr_df2 = pd.DataFrame()
    sup_df1 = pd.DataFrame()
    sup_df2 = pd.DataFrame()
    mrp2_df1 = pd.DataFrame()
    mrp2_df2 = pd.DataFrame()
    ils_df1 = pd.DataFrame()
    ils_df2 = pd.DataFrame()
    iatr_df1 = pd.DataFrame()
    iatr_df2 = pd.DataFrame()
    mdm_df1 = pd.DataFrame()
    mdm_df2 = pd.DataFrame()
    sad_df1 = pd.DataFrame()
    sad_df2 = pd.DataFrame()
    eam_df1 = pd.DataFrame()
    eam_df2 = pd.DataFrame()
 
    reglamenty = pd.DataFrame()
    kommunikazii = pd.DataFrame()
    cody = pd.DataFrame()
    skt = pd.DataFrame()
    obshesistemnoe_po_df1 = pd.DataFrame()
    obshesistemnoe_po_df2 = pd.DataFrame()
    intergracia_oborudovaniya = pd.DataFrame()
    sistemy_monitoringa = pd.DataFrame()
    standarty = pd.DataFrame()
    bi_sistemy = pd.DataFrame()
    ORD = pd.DataFrame()
    kd = pd.DataFrame()
    mzk = pd.DataFrame()
    kadry_1 = pd.DataFrame()
    kadry_2 = pd.DataFrame()
    bim = pd.DataFrame()
    ib = pd.DataFrame()
 
    pattern = r"\s*ИТОГ[ОA0]\s*\(\s*считается\s+[аa@]втоматически,\s+не\s+заполнять\s*\)\s*"
    pattern_ext = r"\s*Итог[оO0]\s*\(\s*заполняется\s+автоматически,\s+не\s+вводить\s+данные\s+вручную\s*\)\s*"
    pattern_amount = r"\s*Количество\s+уникальных\s+наименований\s+отечественного\s+ПО\s*\(\s*считается\s+автоматически,\s+не\s+заполнять\s*\)\s*"
 
    # Вывод информации о загруженных датафреймах
    for key, dataframe in all_sheets_data.items():
        processed_files += 1
        progress = int((processed_files / total_files) * 100)
        progress_callback(progress)
        if key.endswith("1.CAD"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
 
            # Отечественное ПО
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            cad_df1 = pd.concat([cad_df1, df.loc[2:end_index-1].iloc[:, :9]], ignore_index=True)
 
            # Зарубежное ПО
            start_index = df[df.iloc[:, 0] == 'Зарубежное ПО'].index[0] + 1
            end_index = end_indexes[1] if len(end_indexes) > 1 else print(f'не найден индекс итого для {key}')  # Второе появление
            cad_df2 = pd.concat([cad_df2, df.loc[start_index:end_index-1].iloc[:, :9]], ignore_index=True)
 
 
        elif key.endswith("2.ECAD"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
 
            # Отечественное ПО
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            ecad_df1 = pd.concat([ecad_df1, df.loc[2:end_index-1].iloc[:, :9]], ignore_index=True)
 
            # Зарубежное ПО
            start_index = df[df.iloc[:, 0] == 'Зарубежное ПО'].index[0] + 1
            end_index = end_indexes[1] if len(end_indexes) > 1 else print(f'не найден индекс итого для {key}')  # Второе появление
            ecad_df2 = pd.concat([ecad_df2, df.loc[start_index:end_index-1].iloc[:, :9]], ignore_index=True)
 
 
        elif key.endswith("3.CAE"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            # Отечественное ПО
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            cae_df1 = pd.concat([cae_df1, pd.concat([df.loc[2:end_index-1].iloc[:, :9], df.loc[2:end_index-1].iloc[:, 11]], axis=1)], ignore_index=True)
 
            # Зарубежное ПО
            start_index = df[df.iloc[:, 0] == 'Зарубежное ПО'].index[0] + 1
            end_index = end_indexes[1] if len(end_indexes) > 1 else print(f'не найден индекс итого для {key}')  # Второе появление
            cae_df2 = pd.concat([cae_df2, pd.concat([df.loc[start_index:end_index-1].iloc[:, :9], df.loc[start_index:end_index-1].iloc[:, 11]], axis=1)], ignore_index=True)
 
            
        elif key.endswith("4.CAPP"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            # Отечественное ПО
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            capp_df1 = pd.concat([capp_df1, df.loc[2:end_index-1].iloc[:, :9]], ignore_index=True)
 
            # Зарубежное ПО
            start_index = df[df.iloc[:, 0] == 'Зарубежное ПО'].index[0] + 1
            end_index = end_indexes[1] if len(end_indexes) > 1 else print(f'не найден индекс итого для {key}')  # Второе появление
            capp_df2 = pd.concat([capp_df2, df.loc[start_index:end_index-1].iloc[:, :9]], ignore_index=True)
 
 
        elif key.endswith("5.CAM"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            # Отечественное ПО
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'Не найден индекс итого для {key}')  # Первое появление
            cam_df1 = pd.concat([cam_df1, df.loc[2:end_index-1].iloc[:, :9]], ignore_index=True)
 
            # Зарубежное ПО
            start_index = df[df.iloc[:, 0] == 'Зарубежное ПО'].index[0] + 1
            end_index = end_indexes[1] if len(end_indexes) > 1 else print(f'не найден индекс итого для {key}')  # Второе появление
            cam_df2 = pd.concat([cam_df2, df.loc[start_index:end_index-1].iloc[:, :9]], ignore_index=True)
 
 
        elif key.endswith("6.PDM"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            # Отечественное ПО
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            pdm_df1 = pd.concat([pdm_df1, df.loc[2:end_index-1].iloc[:, :9]], ignore_index=True)
 
            # Зарубежное ПО
            start_index = df[df.iloc[:, 0] == 'Зарубежное ПО'].index[0] + 1
            end_index = end_indexes[1] if len(end_indexes) > 1 else print(f'не найден индекс итого для {key}')  # Второе появление
            pdm_df2 = pd.concat([pdm_df2, df.loc[start_index:end_index-1].iloc[:, :9]], ignore_index=True)
 
 
        elif key.endswith("7.ERP"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            # Отечественное ПО
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            erp_df1 = pd.concat([erp_df1, df.loc[2:end_index-1].iloc[:, :9]], ignore_index=True)
 
            # Зарубежное ПО
            start_index = df[df.iloc[:, 0] == 'Зарубежное ПО'].index[0] + 1
            end_index = end_indexes[1] if len(end_indexes) > 1 else print(f'не найден индекс итого для {key}')  # Второе появление
            erp_df2 = pd.concat([erp_df2, df.loc[start_index:end_index-1].iloc[:, :9]], ignore_index=True)
 
 
        elif key.endswith("8.СУБУ"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            # Отечественное ПО
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            subu_df1 = pd.concat([subu_df1, df.loc[2:end_index-1].iloc[:, :9]], ignore_index=True)
 
            # Зарубежное ПО
            start_index = df[df.iloc[:, 0] == 'Зарубежное ПО'].index[0] + 1
            end_index = end_indexes[1] if len(end_indexes) > 1 else print(f'не найден индекс итого для {key}')  # Второе появление
            subu_df2 = pd.concat([subu_df2, df.loc[start_index:end_index-1].iloc[:, :9]], ignore_index=True)
 
 
        elif key.endswith("9.СБ"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            # Отечественное ПО
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            sb_df1 = pd.concat([sb_df1, df.loc[2:end_index-1].iloc[:, :9]], ignore_index=True)
 
            # Зарубежное ПО
            start_index = df[df.iloc[:, 0] == 'Зарубежное ПО'].index[0] + 1
            end_index = end_indexes[1] if len(end_indexes) > 1 else print(f'не найден индекс итого для {key}')  # Второе появление
            sb_df2 = pd.concat([sb_df2, df.loc[start_index:end_index-1].iloc[:, :9]], ignore_index=True)
 
 
        elif key.endswith("10.СУПР"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            # Отечественное ПО
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            supr_df1 = pd.concat([supr_df1, df.loc[2:end_index-1].iloc[:, :9]], ignore_index=True)
 
            # Зарубежное ПО
            start_index = df[df.iloc[:, 0] == 'Зарубежное ПО'].index[0] + 1
            if len(end_indexes)>1:
                end_index = end_indexes[1]
            else:
                end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern_amount, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
                end_index = end_indexes[1] if len(end_indexes) > 1 else print(f'не найден индекс итого для {key}')  # Второе появление
            supr_df2 = pd.concat([supr_df2, df.loc[start_index:end_index-1].iloc[:, :9]], ignore_index=True)
 
 
        elif key.endswith("11.СУП"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            # Отечественное ПО
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            sup_df1 = pd.concat([sup_df1, df.loc[2:end_index-1].iloc[:, :9]], ignore_index=True)
 
            # Зарубежное ПО
            start_index = df[df.iloc[:, 0] == 'Зарубежное ПО'].index[0] + 1
            end_index = end_indexes[1] if len(end_indexes) > 1 else print(f'не найден индекс итого для {key}')  # Второе появление
            sup_df2 = pd.concat([sup_df2, df.loc[start_index:end_index-1].iloc[:, :9]], ignore_index=True)
 
 
        elif key.endswith("12.MRPII"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            # Отечественное ПО
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            mrp2_df1 = pd.concat([mrp2_df1, df.loc[2:end_index-1].iloc[:, :9]], ignore_index=True)
 
            # Зарубежное ПО
            start_index = df[df.iloc[:, 0] == 'Зарубежное ПО'].index[0] + 1
            end_index = end_indexes[1] if len(end_indexes) > 1 else print(f'не найден индекс итого для {key}')  # Второе появление
            mrp2_df2 = pd.concat([mrp2_df2, df.loc[start_index:end_index-1].iloc[:, :9]], ignore_index=True)
 
 
        elif key.endswith("13.ILS"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            # Отечественное ПО
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            ils_df1 = pd.concat([ils_df1, df.loc[2:end_index-1].iloc[:, :9]], ignore_index=True)
 
            # Зарубежное ПО
            start_index = df[df.iloc[:, 0] == 'Зарубежное ПО'].index[0] + 1
            end_index = end_indexes[1] if len(end_indexes) > 1 else print(f'не найден индекс итого для {key}')  # Второе появление
            ils_df2 = pd.concat([ils_df2, df.loc[start_index:end_index-1].iloc[:, :9]], ignore_index=True)
 
 
        elif key.endswith("14.ПО для ИЭТР"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            # Отечественное ПО
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            iatr_df1 = pd.concat([iatr_df1, df.loc[2:end_index-1].iloc[:, :9]], ignore_index=True)
 
            # Зарубежное ПО
            start_index = df[df.iloc[:, 0] == 'Зарубежное ПО'].index[0] + 1
            end_index = end_indexes[1] if len(end_indexes) > 1 else print(f'не найден индекс итого для {key}')  # Второе появление
            iatr_df2 = pd.concat([iatr_df2, df.loc[start_index:end_index-1].iloc[:, :9]], ignore_index=True)
 
 
        elif key.endswith("15.MDM"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            # Отечественное ПО
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            mdm_df1 = pd.concat([mdm_df1, pd.concat([df.loc[2:end_index-1].iloc[:, :11], df.loc[2:end_index-1].iloc[:, 14:18]], axis=1)], ignore_index=True)
 
            # Зарубежное ПО
            start_index = df[df.iloc[:, 0] == 'Зарубежное ПО'].index[0] + 1
            end_index = end_indexes[1] if len(end_indexes) > 1 else print(f'не найден индекс итого для {key}')  # Второе появление
            mdm_df2 = pd.concat([mdm_df2, pd.concat([df.loc[start_index:end_index-1].iloc[:, :11], df.loc[start_index:end_index-1].iloc[:, 14:18]], axis=1)], ignore_index=True)
 
 
        elif key.endswith("16.СЭД"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            # Отечественное ПО
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            sad_df1 = pd.concat([sad_df1, df.loc[2:end_index-1].iloc[:, :9]], ignore_index=True)
 
            # Зарубежное ПО
            start_index = df[df.iloc[:, 0] == 'Зарубежное ПО'].index[0] + 1
            end_index = end_indexes[1] if len(end_indexes) > 1 else print(f'не найден индекс итого для {key}')  # Второе появление
            sad_df2 = pd.concat([sad_df2, df.loc[start_index:end_index-1].iloc[:, :9]], ignore_index=True)
 
 
        elif key.endswith("17.EAM"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            # Отечественное ПО
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            eam_df1 = pd.concat([eam_df1, df.loc[2:end_index-1].iloc[:, :9]], ignore_index=True)
 
            # Зарубежное ПО
            start_index = df[df.iloc[:, 0] == 'Зарубежное ПО'].index[0] + 1
            end_index = end_indexes[1] if len(end_indexes) > 1 else print(f'не найден индекс итого для {key}')  # Второе появление
            eam_df2 = pd.concat([eam_df2, df.loc[start_index:end_index-1].iloc[:, :9]], ignore_index=True)
 
            
        elif key.endswith("18.Регламенты"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            reglamenty = pd.concat([reglamenty, df.loc[1:end_index-1]], ignore_index=True)
 
 
        elif key.endswith("19.Коммуникации"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            kommunikazii = pd.concat([kommunikazii, df.loc[0:5]], ignore_index=True)
            # Добавляем две пустые строки в конец
            kommunikazii = pd.concat([kommunikazii, pd.DataFrame([[None]*len(kommunikazii.columns)])], ignore_index=True)
 
 
        elif key.endswith("20.ЦОДы"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            cody = pd.concat([cody, df.loc[1:end_index-1].iloc[:, :15]], ignore_index=True)
 
 
        elif key.endswith("21.СКТ"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            skt = pd.concat([skt, pd.concat([df.loc[1:end_index-1].iloc[:, :7], df.loc[1:end_index-1].iloc[:, 8:10]], axis=1)], ignore_index=True)
 
 
        elif key.endswith("22.Общесистемное ПО"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
 
            # Отечественное ПО
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            obshesistemnoe_po_df1 = pd.concat([obshesistemnoe_po_df1, df.loc[2:end_index-1].iloc[:, :8]], ignore_index=True)
 
            # Зарубежное ПО
            start_index = df[df.iloc[:, 0] == 'Зарубежное ПО'].index[0] + 1
            end_index = end_indexes[1] if len(end_indexes) > 1 else print(f'не найден индекс итого для {key}')  # Второе появление
            obshesistemnoe_po_df2 = pd.concat([obshesistemnoe_po_df2, df.loc[start_index:end_index-1].iloc[:, :8]], ignore_index=True)
 
 
        elif key.endswith("23. Интеграция оборудования"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_index = df.index[-1]
            intergracia_oborudovaniya = pd.concat([intergracia_oborudovaniya, df.loc[1:end_index].iloc[:, :10]], ignore_index=True)
 
 
        elif key.endswith("24.Системы мониторинга"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern_ext, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            sistemy_monitoringa = pd.concat([sistemy_monitoringa, df.loc[2:end_index-1].iloc[:, :51]], ignore_index=True)
 
 
        elif key.endswith("25. Стандарты"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            standarty = pd.concat([standarty, df.loc[1:end_index-1].iloc[:, :5]], ignore_index=True)
 
 
        elif key.endswith("26.BI-системы"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_index = df.index[-1]
            bi_sistemy = pd.concat([bi_sistemy, df.loc[1:end_index].iloc[:, :11]], ignore_index=True)
 
 
        elif key.endswith("27.ОРД"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            ORD = pd.concat([ORD, df.loc[1:end_index-1].iloc[:, :7]], ignore_index=True)
 
 
        elif key.endswith("28.КД"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            kd = pd.concat([kd, df.loc[1:end_index-1].iloc[:, :10]], ignore_index=True)
 
 
        elif key.endswith("29.МЗК"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            mzk = pd.concat([mzk, df.loc[1:end_index-1].iloc[:, :9]], ignore_index=True)
 
 
        elif key.endswith("30.Кадры 1"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            kadry_1 = pd.concat([kadry_1, df.loc[1:end_index-1].iloc[:, :18]], ignore_index=True)
 
 
        elif key.endswith("31.Кадры 2"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            kadry_2 = pd.concat([kadry_2, df.loc[1:end_index-1].iloc[:, :4]], ignore_index=True)
 
 
        elif key.endswith("32.BIM"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            bim = pd.concat([bim, df.loc[1:end_index-1].iloc[:, :11]], ignore_index=True)
 
 
        elif key.endswith("33.ИБ"):
            print(f"DataFrame '{key}':")
            df = all_sheets_data[key]
            end_indexes = df[df.apply(lambda row: row.astype(str).str.contains(pattern, flags=re.IGNORECASE, regex=True).any(), axis=1)].index
            end_index = end_indexes[0] if len(end_indexes) > 0 else print(f'не найден индекс итого для {key}')  # Первое появление
            ib = pd.concat([ib, df.loc[1:end_index-1].iloc[:, :26]], ignore_index=True)
 
            
        else:
            continue
 
 
    dataframes = [cad_df1, cad_df2,
                  ecad_df1, ecad_df2,
                  cae_df1, cae_df2,
                  capp_df1, capp_df2,
                  cam_df1, cam_df2,
                  pdm_df1, pdm_df2, 
                  erp_df1, erp_df2,
                  subu_df1, subu_df2,
                  sb_df1, sb_df2,
                  supr_df1, supr_df2,
                  sup_df1, sup_df2,
                  mrp2_df1, mrp2_df2,
                  ils_df1, ils_df2,
                  iatr_df1, iatr_df2,
                  mdm_df1, mdm_df2,
                  sad_df1, sad_df2,
                  eam_df1, eam_df2,
                  obshesistemnoe_po_df1, obshesistemnoe_po_df2]
    

    # Перебор каждого датафрейма в списке
    for df in dataframes:
        # Условие, которое проверяет, пустые ли первые четыре столбца
        mask = df.iloc[:, 0:4].isnull().all(axis=1)
        # Удаление строк, где условие истинно
        df.drop(index=df[mask].index, inplace=True)
        # Дополнительно: удаление строк, где все значения пусты
        df.dropna(how='all', inplace=True)

 
    dataframes_base = [reglamenty, cody, skt, intergracia_oborudovaniya, sistemy_monitoringa, standarty, bi_sistemy, ORD, kd, mzk, kadry_1, kadry_2, bim, ib]
 
    for df in dataframes_base:
        # Удаление строк, где первый столбец пуст
        df.dropna(subset=[df.columns[0]], inplace=True)
        # Дополнительно: удаление строк, где все значения пусты
        df.dropna(how='all', inplace=True)
    
    def is_last_column(df, column_name):
        return column_name == df.columns[-1]
    
    def is_last2_column(df, column_name):
        return column_name == df.columns[-2]
    
    def is_last3_column(df, column_name):
        return column_name == df.columns[-3]
    
    def is_last4_column(df, column_name):
        return column_name == df.columns[-4]
    
    def insert_dataframe_to_excel(df, sheet_name, start, limit):
        ws = wb[sheet_name]
        
        # Вставляем значения из датафрейма в листы
        if sheet_name == "3.CAE":
            for column in df:
                for index, v in enumerate(df[column], start=start):
                    if index <= limit+2:
                        if is_last_column(df, column) == True:
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index+1, column=col_number+3, value=v)
                        else:
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index+1, column=col_number+1, value=v)
                    else:
                        for col in ws.iter_cols(min_row=index+1, max_row=index+1, min_col=1, max_col=len(df.columns)):
                            for cell in col:
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        break
        elif sheet_name == "15.MDM":
            for column in df:
                for index, v in enumerate(df[column], start=start):
                    if index <= limit+2:
                        if is_last_column(df, column) == True:
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index+1, column=col_number+4, value=v)
                        elif is_last2_column(df, column) == True:
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index+1, column=col_number+4, value=v)
                        elif is_last3_column(df, column) == True:
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index+1, column=col_number+4, value=v)
                        elif is_last4_column(df, column) == True:
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index+1, column=col_number+4, value=v)
                        else:
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index+1, column=col_number+1, value=v)
                    else:
                        for col in ws.iter_cols(min_row=index+1, max_row=index+1, min_col=1, max_col=len(df.columns)):
                            for cell in col:
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        break
        elif sheet_name == "21.СКТ":
            for column in df:
                for index, v in enumerate(df[column], start=start):
                    if index <= limit+2:
                        if is_last_column(df, column) == True:
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index+1, column=col_number+2, value=v)
                        elif is_last2_column(df, column) == True:
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index+1, column=col_number+2, value=v)
                        else:
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index+1, column=col_number+1, value=v)
                    else:
                        for col in ws.iter_cols(min_row=index+1, max_row=index+1, min_col=1, max_col=len(df.columns)):
                            for cell in col:
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        break
        else:
            for column in df:
                for index, v in enumerate(df[column], start=start):
                    if index <= limit+2:
                        col_number = df.columns.get_loc(column)
                        ws.cell(row=index+1, column=col_number+1, value=v)
                    else:
                        for col in ws.iter_cols(min_row=index+1, max_row=index+1, min_col=1, max_col=len(df.columns)):
                            for cell in col:
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        break
 
        print(f'{sheet_name} - Data inserted successfully')
    
    # # Использовать при сборке финального приложения
    # file_path = getattr(sys, '_MEIPASS', '../../Downloads') + '/DEnew.xlsx'
    # wb = load_workbook(file_path)

    # Использовать при запуске кода
    wb = load_workbook('../assets/DEnew.xlsx')

    # 1.CAD
    insert_dataframe_to_excel(cad_df1, '1.CAD', 3, 2111)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    insert_dataframe_to_excel(cad_df2, '1.CAD', 2118, 1448+2118)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 2.ECAD
    insert_dataframe_to_excel(ecad_df1, '2.ECAD', 3, 365)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    insert_dataframe_to_excel(ecad_df2, '2.ECAD', 372, 439+372)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 3.CAE
    insert_dataframe_to_excel(cae_df1, '3.CAE', 3, 1445)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    insert_dataframe_to_excel(cae_df2, '3.CAE', 1452, 1501+1452)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 4.CAPP
    insert_dataframe_to_excel(capp_df1, '4.CAPP', 3, 754)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    insert_dataframe_to_excel(capp_df2, '4.CAPP', 761, 717+761)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 5.CAM
    insert_dataframe_to_excel(cam_df1, '5.CAM', 3, 1557)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    insert_dataframe_to_excel(cam_df2, '5.CAM', 1564, 1588+1564)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 6.PDM
    insert_dataframe_to_excel(pdm_df1, '6.PDM', 3, 945)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    insert_dataframe_to_excel(pdm_df2, '6.PDM', 952, 649+952)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 7.ERP
    insert_dataframe_to_excel(erp_df1, '7.ERP', 3, 985)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    insert_dataframe_to_excel(erp_df2, '7.ERP', 992, 592+992)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 8.СУБУ
    insert_dataframe_to_excel(subu_df1, '8.СУБУ', 3, 991)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    insert_dataframe_to_excel(subu_df2, '8.СУБУ', 998, 966+998)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 9.СБ
    insert_dataframe_to_excel(sb_df1, '9.СБ', 3, 1088)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    insert_dataframe_to_excel(sb_df2, '9.СБ', 1095, 675+1095)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 10.СУПР
    insert_dataframe_to_excel(supr_df1, '10.СУПР', 3, 807)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    insert_dataframe_to_excel(supr_df2, '10.СУПР', 814, 440+814)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)

    # 11.СУП
    insert_dataframe_to_excel(sup_df1, '11.СУП', 3, 957)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    insert_dataframe_to_excel(sup_df2, '11.СУП', 964, 674+964)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 12.MRPII
    insert_dataframe_to_excel(mrp2_df1, '12.MRPII', 3, 805)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    insert_dataframe_to_excel(mrp2_df2, '12.MRPII', 812, 822+812)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 13.IlS
    insert_dataframe_to_excel(ils_df1, '13.ILS', 3, 690)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    insert_dataframe_to_excel(ils_df2, '13.ILS', 697, 460+697)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 14.ПО для ИЭТР
    insert_dataframe_to_excel(iatr_df1, '14.ПО для ИЭТР', 3, 562)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    insert_dataframe_to_excel(iatr_df2, '14.ПО для ИЭТР', 569, 455+569)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 15.MDM
    insert_dataframe_to_excel(mdm_df1, '15.MDM', 3, 588)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    insert_dataframe_to_excel(mdm_df2, '15.MDM', 595, 739+595)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 16.СЭД
    insert_dataframe_to_excel(sad_df1, '16.СЭД', 3, 614)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    insert_dataframe_to_excel(sad_df2, '16.СЭД', 621, 516+621)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 17.EAM
    insert_dataframe_to_excel(eam_df1, '17.EAM', 3, 498)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    insert_dataframe_to_excel(eam_df2, '17.EAM', 505, 372+505)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
 
    # 18.Регламенты
    insert_dataframe_to_excel(reglamenty, '18.Регламенты', 2, 1726)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 19.Коммуникации
    insert_dataframe_to_excel(kommunikazii, '19.Коммуникации', 7, 3000)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 20.ЦОДы
    insert_dataframe_to_excel(cody, '20.ЦОДы', 2, 983)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 21.СКТ
    insert_dataframe_to_excel(skt, '21.СКТ', 2, 715)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 22.Общесистемное ПО
    insert_dataframe_to_excel(obshesistemnoe_po_df1, '22.Общесистемное ПО', 3, 1563)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    insert_dataframe_to_excel(obshesistemnoe_po_df2, '22.Общесистемное ПО', 1570, 3564+1570)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 23.Интеграция Оборудования
    insert_dataframe_to_excel(intergracia_oborudovaniya, '23. Интеграция оборудования', 2, 1057)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 24.Системы мониторинга
    insert_dataframe_to_excel(sistemy_monitoringa, '24.Системы мониторинга', 3, 644)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 25.Стандарты
    insert_dataframe_to_excel(standarty, '25. Стандарты', 2, 846)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 26.BI-системы
    insert_dataframe_to_excel(bi_sistemy, '26.BI-системы', 2, 647)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 27.ОРД
    insert_dataframe_to_excel(ORD, '27.ОРД', 2, 1108)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 28.КД
    insert_dataframe_to_excel(kd, '28.КД', 2, 649)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 29.МЗК
    insert_dataframe_to_excel(mzk, '29.МЗК', 2, 1081)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 30.Кадры 1
    insert_dataframe_to_excel(kadry_1, '30.Кадры 1', 2, 763)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 31.Кадры 2
    insert_dataframe_to_excel(kadry_2, '31.Кадры 2', 2, 643)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 32.BIM
    insert_dataframe_to_excel(bim, '32.BIM', 2, 776)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)
    
    # 33.ИБ
    insert_dataframe_to_excel(ib, '33.ИБ', 2, 676)
    processed_files += 1
    progress = int((processed_files / total_files) * 100)
    progress_callback(progress)

    wb.save(output_file)
    print('File created successfully')


class ExcelCombinerGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        
        self.setStyleSheet("""
            QMainWindow {
                background-color: #212124;
                color: #ffffff;
            }
            QPushButton {
                color: #ffffff;
                padding-top: 12px;
                padding-bottom: 12px;
                margin-top: 2.5px;
                margin-bottom: 2.5px;
                border-radius: 10px;
                background-color: #161618;
            }
            QListWidget {
                color: #ffffff;
                border-radius: 10px;
                background-color: #000000;
            }
            QLabel {
                color: #ffffff;
            }
            QProgressBar {
                color: #ffffff;
            }
            QMessageBox {
                color: #ffffff;
                background-color: #161618;
            }
        """)

        self.setWindowTitle('Excel Combiner')
        self.setGeometry(100, 100, 800, 600)
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        layout = QVBoxLayout(self.central_widget)

        # Элементы интерфейса
        self.file_list = QListWidget()
        self.add_file_button = QPushButton('Add File')
        self.remove_file_button = QPushButton('Remove File')
        self.combine_button = QPushButton('Merge Files')
        self.status_label = QLabel('Ready for Assembly')
        self.progress_bar = QProgressBar()

        # Добавление элементов в layout
        layout.addWidget(self.file_list)
        layout.addWidget(self.add_file_button)
        layout.addWidget(self.remove_file_button)
        layout.addWidget(self.combine_button)
        layout.addWidget(self.progress_bar)
        layout.addWidget(self.status_label, alignment=Qt.AlignBottom)

        # Привязка событий к кнопкам
        self.add_file_button.clicked.connect(self.add_files)
        self.remove_file_button.clicked.connect(self.remove_files)
        self.combine_button.clicked.connect(self.combine_files)

    def add_files(self):
        file_names, _ = QFileDialog.getOpenFileNames(self, 'Open File', '', 'Excel Files (*.xlsx)')
        self.file_list.addItems(file_names)

    def remove_files(self):
        list_items = self.file_list.selectedItems()
        if not list_items: return
        for item in list_items:
            self.file_list.takeItem(self.file_list.row(item))

    def combine_files(self):
        # Получаем пути ко всем выбранным файлам
        file_paths = [self.file_list.item(i).text() for i in range(self.file_list.count())]
        
        # Указываем путь для сохранения результата
        output_file, _ = QFileDialog.getSaveFileName(self, 'Save File', '', 'Excel Files (*.xlsx)')
        
        if not output_file:
            QMessageBox.warning(self, 'Error', 'You must specify the file to save.')
            return

        try:
            combine_excel_sheets(file_paths, output_file, lambda progress: self.progress_bar.setValue(progress))
            self.status_label.setText('Files merged successfully')
            QMessageBox.information(self, 'Success', 'Files have been merged successfully.')
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'An error occurred while merging files: {e}')
            self.status_label.setText('Error')

def main():
    app = QApplication(sys.argv)
    ex = ExcelCombinerGUI()
    ex.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()

# Команда для сборки, для запуска надо находиться в терминале в папке проекта XLSXAssembler
# pyinstaller --onefile --add-data "DEnew.xlsx;." app.py
# Финальный файл появляется в папке dist, называется app.exe