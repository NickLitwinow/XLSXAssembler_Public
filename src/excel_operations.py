import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


def combine_excel_sheets(file_paths, output_file, progress_callback):
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for idx, file_path in enumerate(file_paths):
            # Load the Excel file
            df = pd.read_excel(file_path)
            df.to_excel(writer, sheet_name=f'Sheet{idx + 1}', index=False)
            progress_callback(int((idx + 1) / len(file_paths) * 100))


def insert_dataframe_to_excel(df, sheet_name, start, limit, wb):
    ws = wb[sheet_name]

    for column in df:
        for index, value in enumerate(df[column], start=start):
            if index <= limit + 2:
                col_number = df.columns.get_loc(column)
                ws.cell(row=index + 1, column=col_number + 1, value=value)
            else:
                # Set empty cells for those beyond the limit
                for col in ws.iter_cols(min_row=index + 1, max_row=index + 1, min_col=1, max_col=len(df.columns)):
                    for cell in col:
                        cell.value = None


def load_workbook_file(file_path):
    return load_workbook(file_path)


# Helper functions
def is_last_column(df, column_name):
    return column_name == df.columns[-1]


def is_last2_column(df, column_name):
    return column_name == df.columns[-2]
