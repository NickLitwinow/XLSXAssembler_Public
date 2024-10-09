from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def insert_excel_data(dfs, output_file):
    dataframes = dfs[0]
    dataframes_base = dfs[1]

    # # Использовать при сборке финального приложения
    # file_path = getattr(sys, '_MEIPASS', '../../Downloads') + '/DEnew.xlsx'
    # wb = load_workbook(file_path)

    # Использовать при запуске кода
    wb = load_workbook('/opt/airflow/pipe/DEnew.xlsx')

    def insert_dataframe_to_excel(df, sheet_name, start, limit):

        def is_last_column(df, column_name):
            return column_name == df.columns[-1]

        def is_last2_column(df, column_name):
            return column_name == df.columns[-2]

        def is_last3_column(df, column_name):
            return column_name == df.columns[-3]

        def is_last4_column(df, column_name):
            return column_name == df.columns[-4]

        ws = wb[sheet_name]

        # Вставляем значения из датафрейма в листы
        if sheet_name == "3.CAE":
            for column in df:
                for index, v in enumerate(df[column], start=start):
                    if index <= limit + 2:
                        if is_last_column(df, column):
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index + 1, column=col_number + 3, value=v)
                        else:
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index + 1, column=col_number + 1, value=v)
                    else:
                        for col in ws.iter_cols(min_row=index + 1, max_row=index + 1, min_col=1,
                                                max_col=len(df.columns)):
                            for cell in col:
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        break
        elif sheet_name == "15.MDM":
            for column in df:
                for index, v in enumerate(df[column], start=start):
                    if index <= limit + 2:
                        if is_last_column(df, column):
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index + 1, column=col_number + 4, value=v)
                        elif is_last2_column(df, column):
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index + 1, column=col_number + 4, value=v)
                        elif is_last3_column(df, column):
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index + 1, column=col_number + 4, value=v)
                        elif is_last4_column(df, column):
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index + 1, column=col_number + 4, value=v)
                        else:
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index + 1, column=col_number + 1, value=v)
                    else:
                        for col in ws.iter_cols(min_row=index + 1, max_row=index + 1, min_col=1,
                                                max_col=len(df.columns)):
                            for cell in col:
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        break
        elif sheet_name == "21.СКТ":
            for column in df:
                for index, v in enumerate(df[column], start=start):
                    if index <= limit + 2:
                        if is_last_column(df, column):
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index + 1, column=col_number + 2, value=v)
                        elif is_last2_column(df, column):
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index + 1, column=col_number + 2, value=v)
                        else:
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index + 1, column=col_number + 1, value=v)
                    else:
                        for col in ws.iter_cols(min_row=index + 1, max_row=index + 1, min_col=1,
                                                max_col=len(df.columns)):
                            for cell in col:
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        break
        else:
            for column in df:
                for index, v in enumerate(df[column], start=start):
                    if index <= limit + 2:
                        col_number = df.columns.get_loc(column)
                        ws.cell(row=index + 1, column=col_number + 1, value=v)
                    else:
                        for col in ws.iter_cols(min_row=index + 1, max_row=index + 1, min_col=1,
                                                max_col=len(df.columns)):
                            for cell in col:
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        break

        print(f'{sheet_name} - Data inserted successfully')

    # 1.CAD
    insert_dataframe_to_excel(dataframes[0], '1.CAD', 3, 2111)
    insert_dataframe_to_excel(dataframes[1], '1.CAD', 2118, 1448 + 2118)

    # 2.ECAD
    insert_dataframe_to_excel(dataframes[2], '2.ECAD', 3, 365)
    insert_dataframe_to_excel(dataframes[3], '2.ECAD', 372, 439 + 372)

    # 3.CAE
    insert_dataframe_to_excel(dataframes[4], '3.CAE', 3, 1445)
    insert_dataframe_to_excel(dataframes[5], '3.CAE', 1452, 1501 + 1452)

    # 4.CAPP
    insert_dataframe_to_excel(dataframes[6], '4.CAPP', 3, 754)
    insert_dataframe_to_excel(dataframes[7], '4.CAPP', 761, 717 + 761)

    # 5.CAM
    insert_dataframe_to_excel(dataframes[8], '5.CAM', 3, 1557)
    insert_dataframe_to_excel(dataframes[9], '5.CAM', 1564, 1588 + 1564)

    # 6.PDM
    insert_dataframe_to_excel(dataframes[10], '6.PDM', 3, 945)
    insert_dataframe_to_excel(dataframes[11], '6.PDM', 952, 649 + 952)

    # 7.ERP
    insert_dataframe_to_excel(dataframes[12], '7.ERP', 3, 985)
    insert_dataframe_to_excel(dataframes[13], '7.ERP', 992, 592 + 992)

    # 8.СУБУ
    insert_dataframe_to_excel(dataframes[14], '8.СУБУ', 3, 991)
    insert_dataframe_to_excel(dataframes[15], '8.СУБУ', 998, 966 + 998)

    # 9.СБ
    insert_dataframe_to_excel(dataframes[16], '9.СБ', 3, 1088)
    insert_dataframe_to_excel(dataframes[17], '9.СБ', 1095, 675 + 1095)

    # 10.СУПР
    insert_dataframe_to_excel(dataframes[18], '10.СУПР', 3, 807)
    insert_dataframe_to_excel(dataframes[19], '10.СУПР', 814, 440 + 814)

    # 11.СУП
    insert_dataframe_to_excel(dataframes[20], '11.СУП', 3, 957)
    insert_dataframe_to_excel(dataframes[21], '11.СУП', 964, 674 + 964)

    # 12.MRPII
    insert_dataframe_to_excel(dataframes[22], '12.MRPII', 3, 805)
    insert_dataframe_to_excel(dataframes[23], '12.MRPII', 812, 822 + 812)

    # 13.IlS
    insert_dataframe_to_excel(dataframes[24], '13.ILS', 3, 690)
    insert_dataframe_to_excel(dataframes[25], '13.ILS', 697, 460 + 697)

    # 14.ПО для ИЭТР
    insert_dataframe_to_excel(dataframes[26], '14.ПО для ИЭТР', 3, 562)
    insert_dataframe_to_excel(dataframes[27], '14.ПО для ИЭТР', 569, 455 + 569)

    # 15.MDM
    insert_dataframe_to_excel(dataframes[28], '15.MDM', 3, 588)
    insert_dataframe_to_excel(dataframes[29], '15.MDM', 595, 739 + 595)

    # 16.СЭД
    insert_dataframe_to_excel(dataframes[30], '16.СЭД', 3, 614)
    insert_dataframe_to_excel(dataframes[31], '16.СЭД', 621, 516 + 621)

    # 17.EAM
    insert_dataframe_to_excel(dataframes[32], '17.EAM', 3, 498)
    insert_dataframe_to_excel(dataframes[33], '17.EAM', 505, 372 + 505)

    # 18.Регламенты
    insert_dataframe_to_excel(dataframes_base[0], '18.Регламенты', 2, 1726)

    # 19.Коммуникации
    insert_dataframe_to_excel(dataframes_base[1], '19.Коммуникации', 7, 3000)

    # 20.ЦОДы
    insert_dataframe_to_excel(dataframes_base[2], '20.ЦОДы', 2, 983)

    # 21.СКТ
    insert_dataframe_to_excel(dataframes_base[3], '21.СКТ', 2, 715)

    # 22.Общесистемное ПО
    insert_dataframe_to_excel(dataframes[34], '22.Общесистемное ПО', 3, 1563)
    insert_dataframe_to_excel(dataframes[35], '22.Общесистемное ПО', 1570, 3564 + 1570)

    # 23.Интеграция Оборудования
    insert_dataframe_to_excel(dataframes_base[4], '23. Интеграция оборудования', 2, 1057)

    # 24.Системы мониторинга
    insert_dataframe_to_excel(dataframes_base[5], '24.Системы мониторинга', 3, 644)

    # 25.Стандарты
    insert_dataframe_to_excel(dataframes_base[6], '25. Стандарты', 2, 846)

    # 26.BI-системы
    insert_dataframe_to_excel(dataframes_base[7], '26.BI-системы', 2, 647)

    # 27.ОРД
    insert_dataframe_to_excel(dataframes_base[8], '27.ОРД', 2, 1108)

    # 28.КД
    insert_dataframe_to_excel(dataframes_base[9], '28.КД', 2, 649)

    # 29.МЗК
    insert_dataframe_to_excel(dataframes_base[10], '29.МЗК', 2, 1081)

    # 30.Кадры 1
    insert_dataframe_to_excel(dataframes_base[11], '30.Кадры 1', 2, 763)

    # 31.Кадры 2
    insert_dataframe_to_excel(dataframes_base[12], '31.Кадры 2', 2, 643)

    # 32.BIM
    insert_dataframe_to_excel(dataframes_base[13], '32.BIM', 2, 776)

    # 33.ИБ
    insert_dataframe_to_excel(dataframes_base[14], '33.ИБ', 2, 676)

    wb.save(output_file)
    print('File created successfully')
