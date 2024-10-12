from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def insert_excel_data(dfs, output_file):
    """
        Inserts data from multiple pandas DataFrames into an Excel workbook, handling
        specific sheet names and applying formatting if necessary.

        Args:
            dfs (list): A list of two lists, where the first contains DataFrames for
                        the main data, and the second contains DataFrames for base data.
            output_file (str): The output file path where the modified Excel file will be saved.

        Functionality:
        - Each DataFrame is inserted into a specific sheet of an existing Excel workbook.
        - Special handling is provided for certain sheets, applying different logic based
          on column positions or sheet names.
        - If data exceeds a specified limit, rows beyond that limit are highlighted in red.
        - Data is inserted into corresponding sheets with special handling for column positions
          depending on the sheet name.

        Workflow:
        - A workbook is loaded, and the DataFrames are inserted one by one into the appropriate
          sheets using the `insert_dataframe_to_excel` function.
        - The workbook is then saved as the output file.

        Notes:
        - This function assumes the workbook `DEnew.xlsx` exists in the specified location
          (i.e., '/opt/airflow/pipe/DEnew.xlsx').
        - It supports complex sheet-specific logic, including color highlighting for cells when
          a row exceeds the given limit.

    """

    dataframes = dfs[0]
    dataframes_base = dfs[1]

    wb = load_workbook('/opt/airflow/pipe/DEnew.xlsx')

    def insert_dataframe_to_excel(df, sheet_name, start, limit):
        """
            Inserts data from a DataFrame into an Excel worksheet and applies specific formatting
            depending on the sheet name.

            Args:
                df (DataFrame): The DataFrame containing the data to be inserted.
                sheet_name (str): The name of the worksheet where the data will be inserted.
                start (int): The row number in the sheet to start inserting data.
                limit (int): The maximum number of rows to insert before applying a red fill
                to excess rows.

            Functionality:
                - Handles special column positioning for certain sheets.
                - Inserts data up to the specified `limit`, beyond which rows are highlighted with
                a red fill (`PatternFill`).
        """

        def is_last_column(df, column_name):
            """Checks if the given column is the last column of the DataFrame."""
            return column_name == df.columns[-1]

        def is_last2_column(df, column_name):
            """Checks if the given column is the second-to-last column of the DataFrame."""
            return column_name == df.columns[-2]

        def is_last3_column(df, column_name):
            """Checks if the given column is the third-to-last column of the DataFrame."""
            return column_name == df.columns[-3]

        def is_last4_column(df, column_name):
            """Checks if the given column is the fourth-to-last column of the DataFrame."""
            return column_name == df.columns[-4]

        ws = wb[sheet_name]

        # Insert values into the worksheet, with different logic for specific sheets
        if sheet_name == "3.CAE":
            for column in df:
                for index, v in enumerate(df[column], start=start):
                    if index <= limit + 2:
                        if is_last_column(df, column):
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index + 1, column=col_number + 3, value=v)
                        else:
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index + 1, column=col_number + 1, value=v)
                    else:
                        for col in ws.iter_cols(min_row=index + 1, max_row=index + 1, min_col=1,
                                                max_col=len(df.columns)):
                            for cell in col:
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        break
        # Handling for sheet "15.MDM"
        elif sheet_name == "15.MDM":
            for column in df:
                for index, v in enumerate(df[column], start=start):
                    if index <= limit + 2:
                        # Different logic for the last four columns
                        if is_last_column(df, column):
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index + 1, column=col_number + 4, value=v)
                        elif is_last2_column(df, column):
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index + 1, column=col_number + 4, value=v)
                        elif is_last3_column(df, column):
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index + 1, column=col_number + 4, value=v)
                        elif is_last4_column(df, column):
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index + 1, column=col_number + 4, value=v)
                        else:
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index + 1, column=col_number + 1, value=v)
                    else:
                        for col in ws.iter_cols(min_row=index + 1, max_row=index + 1, min_col=1,
                                                max_col=len(df.columns)):
                            for cell in col:
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        break
        # Handling for sheet "21.СКТ"
        elif sheet_name == "21.СКТ":
            for column in df:
                for index, v in enumerate(df[column], start=start):
                    if index <= limit + 2:
                        if is_last_column(df, column):
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index + 1, column=col_number + 2, value=v)
                        elif is_last2_column(df, column):
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index + 1, column=col_number + 2, value=v)
                        else:
                            col_number = df.columns.get_loc(column)
                            ws.cell(row=index + 1, column=col_number + 1, value=v)
                    else:
                        for col in ws.iter_cols(min_row=index + 1, max_row=index + 1, min_col=1,
                                                max_col=len(df.columns)):
                            for cell in col:
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        break
        else:
            # Default behavior for all other sheets
            for column in df:
                for index, v in enumerate(df[column], start=start):
                    if index <= limit + 2:
                        col_number = df.columns.get_loc(column)
                        ws.cell(row=index + 1, column=col_number + 1, value=v)
                    else:
                        for col in ws.iter_cols(min_row=index + 1, max_row=index + 1, min_col=1,
                                                max_col=len(df.columns)):
                            for cell in col:
                                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        break

        print(f'{sheet_name} - Data inserted successfully')

    # Calls to insert data into specific sheets
    # 1.CAD, 2.ECAD, 3.CAE, etc.

    # 1.CAD
    insert_dataframe_to_excel(dataframes[0], '1.CAD', 3, 2111)
    insert_dataframe_to_excel(dataframes[1], '1.CAD', 2118, 1448 + 2118)

    # 2.ECAD
    insert_dataframe_to_excel(dataframes[2], '2.ECAD', 3, 365)
    insert_dataframe_to_excel(dataframes[3], '2.ECAD', 372, 439 + 372)

    # 3.CAE
    insert_dataframe_to_excel(dataframes[4], '3.CAE', 3, 1445)
    insert_dataframe_to_excel(dataframes[5], '3.CAE', 1452, 1501 + 1452)

    # 4.CAPP
    insert_dataframe_to_excel(dataframes[6], '4.CAPP', 3, 754)
    insert_dataframe_to_excel(dataframes[7], '4.CAPP', 761, 717 + 761)

    # 5.CAM
    insert_dataframe_to_excel(dataframes[8], '5.CAM', 3, 1557)
    insert_dataframe_to_excel(dataframes[9], '5.CAM', 1564, 1588 + 1564)

    # 6.PDM
    insert_dataframe_to_excel(dataframes[10], '6.PDM', 3, 945)
    insert_dataframe_to_excel(dataframes[11], '6.PDM', 952, 649 + 952)

    # 7.ERP
    insert_dataframe_to_excel(dataframes[12], '7.ERP', 3, 985)
    insert_dataframe_to_excel(dataframes[13], '7.ERP', 992, 592 + 992)

    # 8.СУБУ
    insert_dataframe_to_excel(dataframes[14], '8.СУБУ', 3, 991)
    insert_dataframe_to_excel(dataframes[15], '8.СУБУ', 998, 966 + 998)

    # 9.СБ
    insert_dataframe_to_excel(dataframes[16], '9.СБ', 3, 1088)
    insert_dataframe_to_excel(dataframes[17], '9.СБ', 1095, 675 + 1095)

    # 10.СУПР
    insert_dataframe_to_excel(dataframes[18], '10.СУПР', 3, 807)
    insert_dataframe_to_excel(dataframes[19], '10.СУПР', 814, 440 + 814)

    # 11.СУП
    insert_dataframe_to_excel(dataframes[20], '11.СУП', 3, 957)
    insert_dataframe_to_excel(dataframes[21], '11.СУП', 964, 674 + 964)

    # 12.MRPII
    insert_dataframe_to_excel(dataframes[22], '12.MRPII', 3, 805)
    insert_dataframe_to_excel(dataframes[23], '12.MRPII', 812, 822 + 812)

    # 13.IlS
    insert_dataframe_to_excel(dataframes[24], '13.ILS', 3, 690)
    insert_dataframe_to_excel(dataframes[25], '13.ILS', 697, 460 + 697)

    # 14.ПО для ИЭТР
    insert_dataframe_to_excel(dataframes[26], '14.ПО для ИЭТР', 3, 562)
    insert_dataframe_to_excel(dataframes[27], '14.ПО для ИЭТР', 569, 455 + 569)

    # 15.MDM
    insert_dataframe_to_excel(dataframes[28], '15.MDM', 3, 588)
    insert_dataframe_to_excel(dataframes[29], '15.MDM', 595, 739 + 595)

    # 16.СЭД
    insert_dataframe_to_excel(dataframes[30], '16.СЭД', 3, 614)
    insert_dataframe_to_excel(dataframes[31], '16.СЭД', 621, 516 + 621)

    # 17.EAM
    insert_dataframe_to_excel(dataframes[32], '17.EAM', 3, 498)
    insert_dataframe_to_excel(dataframes[33], '17.EAM', 505, 372 + 505)

    # 18.Регламенты
    insert_dataframe_to_excel(dataframes_base[0], '18.Регламенты', 2, 1726)

    # 19.Коммуникации
    insert_dataframe_to_excel(dataframes_base[1], '19.Коммуникации', 7, 3000)

    # 20.ЦОДы
    insert_dataframe_to_excel(dataframes_base[2], '20.ЦОДы', 2, 983)

    # 21.СКТ
    insert_dataframe_to_excel(dataframes_base[3], '21.СКТ', 2, 715)

    # 22.Общесистемное ПО
    insert_dataframe_to_excel(dataframes[34], '22.Общесистемное ПО', 3, 1563)
    insert_dataframe_to_excel(dataframes[35], '22.Общесистемное ПО', 1570, 3564 + 1570)

    # 23.Интеграция Оборудования
    insert_dataframe_to_excel(dataframes_base[4], '23. Интеграция оборудования', 2, 1057)

    # 24.Системы мониторинга
    insert_dataframe_to_excel(dataframes_base[5], '24.Системы мониторинга', 3, 644)

    # 25.Стандарты
    insert_dataframe_to_excel(dataframes_base[6], '25. Стандарты', 2, 846)

    # 26.BI-системы
    insert_dataframe_to_excel(dataframes_base[7], '26.BI-системы', 2, 647)

    # 27.ОРД
    insert_dataframe_to_excel(dataframes_base[8], '27.ОРД', 2, 1108)

    # 28.КД
    insert_dataframe_to_excel(dataframes_base[9], '28.КД', 2, 649)

    # 29.МЗК
    insert_dataframe_to_excel(dataframes_base[10], '29.МЗК', 2, 1081)

    # 30.Кадры 1
    insert_dataframe_to_excel(dataframes_base[11], '30.Кадры 1', 2, 763)

    # 31.Кадры 2
    insert_dataframe_to_excel(dataframes_base[12], '31.Кадры 2', 2, 643)

    # 32.BIM
    insert_dataframe_to_excel(dataframes_base[13], '32.BIM', 2, 776)

    # 33.ИБ
    insert_dataframe_to_excel(dataframes_base[14], '33.ИБ', 2, 676)

    wb.save(output_file)
    print('File created successfully')
